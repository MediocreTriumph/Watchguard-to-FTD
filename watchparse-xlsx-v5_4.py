#!/usr/bin/env python3
"""
WatchGuard Parser v5 - Excel Workbook Output

Parses WatchGuard XML configuration and outputs structured data to an Excel
workbook with separate worksheets per object category. Same parsing logic as
watchparse-json-v5.py but targets .xlsx instead of .json.

Worksheets:
  - Summary          Overview counts and metadata
  - Hosts            Host address objects
  - Networks         Network address objects
  - Ranges           Address range objects
  - FQDNs            FQDN address objects
  - Address Groups   Alias-based address groups
  - Interface Aliases
  - TCP Services
  - UDP Services
  - ICMP Services
  - Protocol Only    GRE, ESP, etc.
  - Other Services   Unsupported protocols
  - Service Groups
  - Policies
  - NAT Rules
  - Routes
  - Interfaces
  - SDWAN Actions
  - App Actions
  - Geo Actions
  - Validation       Reference validation issues
"""

import sys
import xml.etree.ElementTree as ET
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================================
# Protocol mappings (identical to JSON version)
# ============================================================================

PROTOCOLS = {
    "0": "HOPOPT", "1": "ICMP", "2": "IGMP", "6": "TCP",
    "17": "UDP", "47": "GRE", "50": "ESP", "51": "AH",
    "58": "ICMPv6", "89": "OSPFIGP"
}

PORT_BASED_PROTOCOLS = {"TCP", "UDP"}
ICMP_PROTOCOLS = {"ICMP", "ICMPv6"}
PROTOCOL_ONLY = {"GRE", "ESP", "AH", "IGMP", "OSPFIGP"}
UNSUPPORTED_PROTOCOLS = {"HOPOPT"}


# ============================================================================
# Excel formatting constants
# ============================================================================

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)

DATA_FONT = Font(name="Arial", size=10)
DATA_ALIGNMENT = Alignment(horizontal="left", vertical="top", wrap_text=True)

SUMMARY_LABEL_FONT = Font(name="Arial", bold=True, size=10)
SUMMARY_VALUE_FONT = Font(name="Arial", size=10, color="2F5496")
SUMMARY_SECTION_FONT = Font(name="Arial", bold=True, size=12, color="2F5496")
SUMMARY_SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")

WARNING_FONT = Font(name="Arial", size=10, color="CC0000")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")

THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9")
)


def style_header_row(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}1"


def style_data_rows(ws, start_row, end_row, num_cols):
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER


def auto_width(ws, num_cols, max_width=50, min_width=10):
    for col in range(1, num_cols + 1):
        max_len = min_width
        letter = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            cell = row[0]
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[letter].width = max_len + 2


def write_sheet(ws, headers, rows):
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    num_cols = len(headers)
    style_header_row(ws, num_cols)
    if rows:
        style_data_rows(ws, 2, len(rows) + 1, num_cols)
    auto_width(ws, num_cols)


# ============================================================================
# Service name helpers (identical to JSON version)
# ============================================================================

def sanitize_port_for_name(port):
    if not port:
        return ""
    return port.replace(",", "_")


def generate_service_name(original_name, protocol, port=None):
    if port:
        return f"{original_name}_{protocol}_{sanitize_port_for_name(port)}"
    return f"{original_name}_{protocol}"


# ============================================================================
# Main parser (same logic as JSON version, returns structured dicts)
# ============================================================================

def parse_watchguard_config(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    config = {
        "metadata": {
            "parsed_date": datetime.now().isoformat(),
            "source_file": xml_file,
            "parser_version": "v5-xlsx"
        },
        "addresses": {"hosts": {}, "networks": {}, "ranges": {}, "fqdns": {}},
        "address_groups": [],
        "interface_aliases": [],
        "services": {"tcp": [], "udp": [], "icmp": [], "protocol_only": [], "other": []},
        "service_groups": [],
        "routes": [],
        "interfaces": [],
        "policies": [],
        "nat_rules": [],
        "sdwan_actions": [],
        "app_actions": [],
        "geo_actions": []
    }

    # === PHASE 1: Collect all service definitions ===
    services_by_name = defaultdict(list)

    for service in root.findall("./service-list/service"):
        name = ""
        description = ""
        for elem in service:
            if elem.tag == "name":
                name = elem.text
            elif elem.tag == "description":
                description = elem.text if elem.text else ""
            elif elem.tag == "service-item":
                for item in elem:
                    protocol = None
                    port = None
                    for field in item:
                        if field.tag == "protocol":
                            protocol = PROTOCOLS.get(field.text, field.text)
                        elif field.tag == "server-port":
                            port = field.text
                    if protocol and name:
                        services_by_name[name].append({
                            "protocol": protocol, "port": port, "description": description
                        })

    # === PHASE 2: Process services, create groups for duplicates ===
    for original_name, service_list in services_by_name.items():
        needs_group = len(service_list) > 1
        tcp_udp_members = []
        icmp_members = []
        protocol_members = []
        warnings = []

        for svc in service_list:
            protocol = svc["protocol"]
            port = svc["port"]
            description = svc["description"]

            unique_name = generate_service_name(original_name, protocol, port) if needs_group else original_name

            service_obj = {
                "name": unique_name,
                "original_name": original_name,
                "description": f"{original_name} service ({protocol}" + (f"/{port})" if port else ")"),
            }
            if port:
                service_obj["port"] = port

            if protocol in PORT_BASED_PROTOCOLS:
                if protocol == "TCP":
                    config["services"]["tcp"].append(service_obj)
                else:
                    config["services"]["udp"].append(service_obj)
                if needs_group:
                    tcp_udp_members.append(unique_name)

            elif protocol in ICMP_PROTOCOLS:
                service_obj["icmp_version"] = "v6" if protocol == "ICMPv6" else "v4"
                config["services"]["icmp"].append(service_obj)
                if needs_group:
                    icmp_members.append(unique_name)

            elif protocol in PROTOCOL_ONLY:
                service_obj["protocol"] = protocol
                service_obj["protocol_number"] = {v: k for k, v in PROTOCOLS.items()}.get(protocol, "")
                config["services"]["protocol_only"].append(service_obj)
                if needs_group:
                    protocol_members.append(unique_name)

            elif protocol in UNSUPPORTED_PROTOCOLS:
                service_obj["protocol"] = protocol
                config["services"]["other"].append(service_obj)
                warnings.append(f"Unsupported protocol {protocol} skipped")

            else:
                service_obj["protocol"] = protocol
                config["services"]["other"].append(service_obj)
                warnings.append(f"Unknown protocol {protocol} - may require manual handling")

        if needs_group:
            group = {
                "name": f"{original_name}_svc_group",
                "original_name": original_name,
                "description": f"Service group for {original_name} ({len(service_list)} original services)",
                "members": tcp_udp_members,
            }
            if icmp_members:
                group["icmp_members"] = icmp_members
            if protocol_members:
                group["protocol_members"] = protocol_members
            if warnings:
                group["warnings"] = list(set(warnings))
            config["service_groups"].append(group)

    # === Parse Routes ===
    for route in root.findall("./system-parameters/route/route-entry"):
        route_obj = {}
        for elem in route:
            if elem.tag == "dest-address":
                route_obj["destination"] = elem.text
            elif elem.tag == "mask":
                route_obj["mask"] = elem.text
            elif elem.tag == "gateway-ip":
                route_obj["gateway"] = elem.text
        if route_obj:
            config["routes"].append(route_obj)

    # === Parse Address Objects ===
    for addr_group in root.findall("./address-group-list/address-group"):
        name = ""
        description = ""
        for elem in addr_group:
            if elem.tag == "name":
                name = elem.text
            elif elem.tag == "description":
                description = elem.text if elem.text else ""
            elif elem.tag == "addr-group-member":
                for member in elem:
                    host_ip = network_addr = mask = start_addr = end_addr = fqdn = None
                    for field in member:
                        if field.tag == "host-ip-addr":
                            host_ip = field.text
                        elif field.tag == "ip-network-addr":
                            network_addr = field.text
                        elif field.tag == "ip-mask":
                            mask = field.text
                        elif field.tag == "start-ip-addr":
                            start_addr = field.text
                        elif field.tag == "end-ip-addr":
                            end_addr = field.text
                        elif field.tag == "domain":
                            fqdn = field.text

                    if host_ip:
                        if name not in config["addresses"]["hosts"]:
                            config["addresses"]["hosts"][name] = {"name": name, "description": description, "ip": host_ip}
                    elif network_addr and mask:
                        if name not in config["addresses"]["networks"]:
                            config["addresses"]["networks"][name] = {"name": name, "description": description, "network": network_addr, "mask": mask}
                    elif start_addr and end_addr:
                        if name not in config["addresses"]["ranges"]:
                            config["addresses"]["ranges"][name] = {"name": name, "description": description, "start": start_addr, "end": end_addr}
                    elif fqdn:
                        if name not in config["addresses"]["fqdns"]:
                            config["addresses"]["fqdns"][name] = {"name": name, "description": description, "fqdn": fqdn}

    # === Parse Aliases ===
    for alias in root.findall("./alias-list/alias"):
        alias_obj = {
            "name": "", "description": "",
            "member_types": [], "member_users": [], "members": [],
            "alias_references": [], "member_interfaces": []
        }
        for elem in alias:
            if elem.tag == "name":
                alias_obj["name"] = elem.text
            elif elem.tag == "description":
                alias_obj["description"] = elem.text if elem.text else ""
            elif elem.tag == "alias-member-list":
                for alias_member in elem.findall("alias-member"):
                    for field in alias_member:
                        if field.tag == "type":
                            alias_obj["member_types"].append(field.text)
                        elif field.tag == "user":
                            alias_obj["member_users"].append(field.text)
                        elif field.tag == "address":
                            alias_obj["members"].append(field.text)
                        elif field.tag == "alias-name":
                            alias_obj["alias_references"].append(field.text)
                        elif field.tag == "interface":
                            alias_obj["member_interfaces"].append(field.text)

        alias_obj["member_types"] = list(dict.fromkeys(alias_obj["member_types"]))
        alias_obj["member_users"] = list(dict.fromkeys(alias_obj["member_users"]))
        alias_obj["members"] = list(dict.fromkeys(alias_obj["members"]))
        alias_obj["alias_references"] = list(dict.fromkeys(alias_obj["alias_references"]))
        alias_obj["member_interfaces"] = list(dict.fromkeys(alias_obj["member_interfaces"]))

        if alias_obj["members"] and all(m == "Any" for m in alias_obj["members"]):
            alias_obj["members"] = ["Any"]

        is_interface_alias = (
            len(alias_obj["member_interfaces"]) > 0
            and len(alias_obj["members"]) == 0
            and len(alias_obj["alias_references"]) == 0
        ) or alias_obj["name"] in ["Any", "Firebox", "Any-External", "Any-Trusted", "Any-Optional"]

        if is_interface_alias:
            config["interface_aliases"].append(alias_obj)
        else:
            config["address_groups"].append(alias_obj)

    # === Parse Interfaces ===
    for interface in root.findall("./interface-list/interface"):
        intf = {
            "name": "", "description": "", "device_name": "", "enabled": "",
            "node_type": "", "ip": "", "gateway": "", "mask": "", "secondary_ips": []
        }
        for elem in interface:
            if elem.tag == "name":
                intf["name"] = elem.text
            elif elem.tag == "description":
                intf["description"] = elem.text if elem.text else ""
            elif elem.tag == "if-item-list":
                for item in elem:
                    for physif in item:
                        for field in physif:
                            if field.tag == "if-dev-name":
                                intf["device_name"] = field.text
                            elif field.tag == "enabled":
                                intf["enabled"] = field.text
                            elif field.tag == "ip-node-type":
                                intf["node_type"] = field.text
                            elif field.tag == "ip":
                                intf["ip"] = field.text
                            elif field.tag == "default-gateway":
                                intf["gateway"] = field.text
                            elif field.tag == "netmask":
                                intf["mask"] = field.text
                            elif field.tag == "secondary-ip-list":
                                for sec_ip in field:
                                    for ip_elem in sec_ip:
                                        if ip_elem.text:
                                            intf["secondary_ips"].append(ip_elem.text)
        config["interfaces"].append(intf)

    # === Build alias lookup for policy resolution ===
    alias_map = {}
    for a in config["address_groups"] + config["interface_aliases"]:
        alias_map[a["name"]] = a

    def resolve_alias_members(alias_name, visited=None):
        if visited is None:
            visited = set()
        if alias_name in visited:
            return []
        visited.add(alias_name)
        if alias_name == "Any":
            return ["Any"]
        if alias_name not in alias_map:
            return []
        alias = alias_map[alias_name]
        resolved = list(alias["members"])
        for ref in alias["alias_references"]:
            resolved.extend(resolve_alias_members(ref, visited))
        return resolved

    # === Parse Policies ===
    for policy in root.findall("./abs-policy-list/abs-policy"):
        pol = {
            "name": "", "source_aliases": [], "destination_aliases": [],
            "source_members": [], "destination_members": [],
            "service": "", "enabled": "", "action": "", "nat_policy": "",
            "description": "", "reject_action": "", "tag": "", "schedule": "",
            "log_enabled": "", "route_policy": "", "proxy": "",
            "sdwan_action": "", "app_action": ""
        }
        for elem in policy:
            if elem.tag == "name":
                pol["name"] = elem.text
            elif elem.tag == "from-alias-list":
                for alias_ref in elem:
                    alias_name = alias_ref.text
                    pol["source_aliases"].append(alias_name)
                    pol["source_members"].extend(resolve_alias_members(alias_name))
            elif elem.tag == "to-alias-list":
                for alias_ref in elem:
                    alias_name = alias_ref.text
                    pol["destination_aliases"].append(alias_name)
                    pol["destination_members"].extend(resolve_alias_members(alias_name))
            elif elem.tag == "service":
                pol["service"] = elem.text
            elif elem.tag == "enabled":
                pol["enabled"] = elem.text
            elif elem.tag == "firewall":
                pol["action"] = elem.text
            elif elem.tag == "policy-nat":
                pol["nat_policy"] = elem.text
            elif elem.tag == "description":
                pol["description"] = elem.text if elem.text else ""
            elif elem.tag == "reject-action":
                pol["reject_action"] = elem.text
            elif elem.tag == "tag-list":
                pol["tag"] = ",".join([a.text for a in elem if a.text])
            elif elem.tag == "settings":
                for setting in elem:
                    if setting.tag == "schedule":
                        pol["schedule"] = setting.text
                    elif setting.tag == "log-enabled":
                        pol["log_enabled"] = setting.text
                    elif setting.tag == "policy-routing":
                        pol["route_policy"] = setting.text
                    elif setting.tag == "proxy":
                        pol["proxy"] = setting.text
                    elif setting.tag == "sdwan-action":
                        pol["sdwan_action"] = setting.text
            elif elem.tag == "app-action":
                pol["app_action"] = elem.text if elem.text else ""

        pol["source_members"] = list(dict.fromkeys(pol["source_members"]))
        pol["destination_members"] = list(dict.fromkeys(pol["destination_members"]))
        config["policies"].append(pol)

    # === Parse NAT Rules ===
    for nat in root.findall("./nat-list/nat"):
        nat_rule = {
            "name": "", "type": "", "algorithm": "", "proxy_arp": "",
            "address_type": "", "port": "", "external_address": "",
            "interface": "", "internal_address": ""
        }
        for elem in nat:
            if elem.tag == "name":
                nat_rule["name"] = elem.text
            elif elem.tag == "type":
                nat_rule["type"] = elem.text
            elif elem.tag == "algorithm":
                nat_rule["algorithm"] = elem.text
            elif elem.tag == "proxy-arp":
                nat_rule["proxy_arp"] = elem.text
            elif elem.tag == "nat-item":
                for item in elem:
                    for field in item:
                        if field.tag == "addr-type":
                            nat_rule["address_type"] = field.text
                        elif field.tag == "port":
                            nat_rule["port"] = field.text
                        elif field.tag == "ext-addr-name":
                            nat_rule["external_address"] = field.text
                        elif field.tag == "interface":
                            nat_rule["interface"] = field.text
                        elif field.tag == "addr-name":
                            nat_rule["internal_address"] = field.text
        config["nat_rules"].append(nat_rule)

    # === Parse SDWAN Actions ===
    SDWAN_ALGORITHMS = {"1": "Global", "2": "Failover (Immediate Failback)"}
    for sdwan in root.findall("./sdwan-action-list/sdwan-action"):
        action = {
            "name": "", "description": "", "algorithm": "",
            "algorithm_description": "", "interfaces": [],
            "primary_interface": "", "secondary_interface": "",
            "failback_grace_period": ""
        }
        for elem in sdwan:
            if elem.tag == "name":
                action["name"] = elem.text
            elif elem.tag == "description":
                action["description"] = elem.text if elem.text else ""
            elif elem.tag == "algorithm":
                action["algorithm"] = elem.text
                action["algorithm_description"] = SDWAN_ALGORITHMS.get(elem.text, f"Unknown ({elem.text})")
            elif elem.tag == "failback-grace-period":
                action["failback_grace_period"] = elem.text
            elif elem.tag == "if-list":
                for if_name in elem.findall("if-name"):
                    action["interfaces"].append(if_name.text)
                if len(action["interfaces"]) >= 1:
                    action["primary_interface"] = action["interfaces"][0]
                if len(action["interfaces"]) >= 2:
                    action["secondary_interface"] = action["interfaces"][1]
        config["sdwan_actions"].append(action)

    # === Parse Application Control ===
    for app in root.findall("./app-action-list/app-action"):
        app_action = {
            "name": "", "description": "",
            "allowed_apps": [], "blocked_apps": [], "fallthrough_action": ""
        }
        for elem in app:
            if elem.tag == "name":
                app_action["name"] = elem.text
            elif elem.tag == "description":
                app_action["description"] = elem.text if elem.text else ""
            elif elem.tag == "fallthrough":
                app_action["fallthrough_action"] = elem.text
            elif elem.tag == "allow-list":
                for app_elem in elem.findall("app"):
                    for field in app_elem:
                        if field.tag == "name":
                            app_action["allowed_apps"].append(field.text)
            elif elem.tag == "block-list":
                for app_elem in elem.findall("app"):
                    for field in app_elem:
                        if field.tag == "name":
                            app_action["blocked_apps"].append(field.text)
        config["app_actions"].append(app_action)

    # === Parse Geo Blocking ===
    for geo in root.findall(".//geo-action-list/geo-action"):
        geo_action = {"name": "", "description": "", "blocked_countries": []}
        for elem in geo:
            if elem.tag == "name":
                geo_action["name"] = elem.text
            elif elem.tag == "description":
                geo_action["description"] = elem.text if elem.text else ""
            elif elem.tag == "geo-list":
                for geo_elem in elem.findall("geo"):
                    for field in geo_elem:
                        if field.tag == "country":
                            geo_action["blocked_countries"].append(field.text)
        config["geo_actions"].append(geo_action)

    # === Convert address dicts to lists ===
    config["addresses"]["hosts"] = list(config["addresses"]["hosts"].values())
    config["addresses"]["networks"] = list(config["addresses"]["networks"].values())
    config["addresses"]["ranges"] = list(config["addresses"]["ranges"].values())
    config["addresses"]["fqdns"] = list(config["addresses"]["fqdns"].values())

    return config


# ============================================================================
# Validation (same as JSON version)
# ============================================================================

def validate_references(config):
    all_addresses = set()
    for addr_type in ["hosts", "networks", "ranges", "fqdns"]:
        for obj in config["addresses"][addr_type]:
            all_addresses.add(obj["name"])
    for alias in config["interface_aliases"]:
        all_addresses.add(alias["name"])

    all_group_names = set()
    for group in config["address_groups"]:
        all_group_names.add(group["name"])

    issues = {"broken_address_references": [], "broken_alias_references": []}

    for group in config["address_groups"]:
        for member in group["members"]:
            if member != "Any" and member not in all_addresses and member not in all_group_names:
                issues["broken_address_references"].append({
                    "group": group["name"], "missing_member": member
                })
        for ref in group["alias_references"]:
            if ref not in all_group_names and ref not in [a["name"] for a in config["interface_aliases"]]:
                issues["broken_alias_references"].append({
                    "group": group["name"], "missing_alias": ref
                })
    return issues


# ============================================================================
# Excel workbook writer
# ============================================================================

def build_address_resolver(config):
    """
    Build a comprehensive resolver that can chase any name -- whether it's a
    direct address object, an alias, or an address group -- all the way down
    to actual IP/network/range/FQDN values.

    Returns a function: resolve(name) -> list of value strings

    Resolution chain:
      1. Direct address object (host/network/range/fqdn) -> value immediately
      2. Alias or address group -> recurse into members + alias_references
      3. "Any" -> ["Any"]
      4. Unresolved -> ["[!] <name>"]
    """
    # Layer 1: direct address objects (leaf nodes with real values)
    direct_lookup = {}
    for h in config["addresses"]["hosts"]:
        direct_lookup[h["name"]] = [h["ip"]]
    for n in config["addresses"]["networks"]:
        direct_lookup[n["name"]] = [f"{n['network']}/{n['mask']}"]
    for r in config["addresses"]["ranges"]:
        direct_lookup[r["name"]] = [f"{r['start']} - {r['end']}"]
    for f in config["addresses"]["fqdns"]:
        direct_lookup[f["name"]] = [f["fqdn"]]

    # Layer 2: alias/group map (intermediate nodes that reference other names)
    alias_map = {}
    for grp in config["address_groups"]:
        alias_map[grp["name"]] = grp
    for ifa in config["interface_aliases"]:
        alias_map[ifa["name"]] = ifa

    def resolve(name, visited=None):
        if visited is None:
            visited = set()
        if name in visited:
            return []
        visited.add(name)

        if name == "Any":
            return ["Any"]

        # Check direct address objects first (leaf resolution)
        if name in direct_lookup:
            return list(direct_lookup[name])

        # Check alias/group -- recurse into members and alias_references
        if name in alias_map:
            alias = alias_map[name]
            values = []
            for member in alias.get("members", []):
                values.extend(resolve(member, visited.copy()))
            for ref in alias.get("alias_references", []):
                values.extend(resolve(ref, visited.copy()))
            if values:
                return values

        # Unresolvable
        return [f"[!] {name}"]

    return resolve


def resolve_policy_values(alias_names, resolver):
    """
    Resolve a policy's source or destination alias list to actual values.
    Starts from the alias names (not pre-resolved members) so we chase the
    full chain: alias -> group -> members -> address objects -> values.
    """
    all_values = []
    for alias_name in alias_names:
        all_values.extend(resolver(alias_name))
    # Deduplicate while preserving order
    return list(dict.fromkeys(all_values))


def write_workbook(config, issues, output_file):
    wb = Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "2F5496"

    meta = config["metadata"]
    summary_data = [
        ("METADATA", None),
        ("Source File", meta["source_file"]),
        ("Parsed Date", meta["parsed_date"]),
        ("Parser Version", meta["parser_version"]),
        ("", ""),
        ("ADDRESS OBJECTS", None),
        ("Hosts", len(config["addresses"]["hosts"])),
        ("Networks", len(config["addresses"]["networks"])),
        ("Ranges", len(config["addresses"]["ranges"])),
        ("FQDNs", len(config["addresses"]["fqdns"])),
        ("Address Groups", len(config["address_groups"])),
        ("Interface Aliases", len(config["interface_aliases"])),
        ("", ""),
        ("SERVICE OBJECTS", None),
        ("TCP Services", len(config["services"]["tcp"])),
        ("UDP Services", len(config["services"]["udp"])),
        ("ICMP Services", len(config["services"]["icmp"])),
        ("Protocol-Only (GRE, ESP, etc.)", len(config["services"]["protocol_only"])),
        ("Other/Unsupported", len(config["services"]["other"])),
        ("Service Groups", len(config["service_groups"])),
        ("", ""),
        ("POLICIES & OTHER", None),
        ("Policies", len(config["policies"])),
        ("NAT Rules", len(config["nat_rules"])),
        ("Routes", len(config["routes"])),
        ("Interfaces", len(config["interfaces"])),
        ("SDWAN Actions", len(config["sdwan_actions"])),
        ("App Actions", len(config["app_actions"])),
        ("Geo Actions", len(config["geo_actions"])),
        ("", ""),
        ("VALIDATION", None),
        ("Broken Address References", len(issues["broken_address_references"])),
        ("Broken Alias References", len(issues["broken_alias_references"])),
    ]

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 50

    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell_a = ws.cell(row=row_idx, column=1, value=label)
        if value is None:
            # Section header
            cell_a.font = SUMMARY_SECTION_FONT
            cell_a.fill = SUMMARY_SECTION_FILL
            ws.cell(row=row_idx, column=2).fill = SUMMARY_SECTION_FILL
        else:
            cell_a.font = SUMMARY_LABEL_FONT
            cell_b = ws.cell(row=row_idx, column=2, value=value)
            cell_b.font = SUMMARY_VALUE_FONT

    # --- Hosts ---
    ws_hosts = wb.create_sheet("Hosts")
    write_sheet(ws_hosts,
        ["Name", "Description", "IP Address"],
        [(h["name"], h["description"], h["ip"]) for h in config["addresses"]["hosts"]]
    )

    # --- Networks ---
    ws_nets = wb.create_sheet("Networks")
    write_sheet(ws_nets,
        ["Name", "Description", "Network", "Mask"],
        [(n["name"], n["description"], n["network"], n["mask"]) for n in config["addresses"]["networks"]]
    )

    # --- Ranges ---
    ws_ranges = wb.create_sheet("Ranges")
    write_sheet(ws_ranges,
        ["Name", "Description", "Start IP", "End IP"],
        [(r["name"], r["description"], r["start"], r["end"]) for r in config["addresses"]["ranges"]]
    )

    # --- FQDNs ---
    ws_fqdns = wb.create_sheet("FQDNs")
    write_sheet(ws_fqdns,
        ["Name", "Description", "FQDN"],
        [(f["name"], f["description"], f["fqdn"]) for f in config["addresses"]["fqdns"]]
    )

    # --- Address Groups ---
    ws_ag = wb.create_sheet("Address Groups")
    write_sheet(ws_ag,
        ["Name", "Description", "Member Types", "Members", "Alias References", "Member Users", "Member Interfaces"],
        [
            (
                g["name"], g["description"],
                "; ".join(g["member_types"]),
                "; ".join(g["members"]),
                "; ".join(g["alias_references"]),
                "; ".join(g["member_users"]),
                "; ".join(g["member_interfaces"])
            )
            for g in config["address_groups"]
        ]
    )

    # --- Interface Aliases ---
    ws_ia = wb.create_sheet("Interface Aliases")
    write_sheet(ws_ia,
        ["Name", "Description", "Member Types", "Member Interfaces", "Members", "Alias References"],
        [
            (
                a["name"], a["description"],
                "; ".join(a["member_types"]),
                "; ".join(a["member_interfaces"]),
                "; ".join(a["members"]),
                "; ".join(a["alias_references"])
            )
            for a in config["interface_aliases"]
        ]
    )

    # --- TCP Services ---
    ws_tcp = wb.create_sheet("TCP Services")
    write_sheet(ws_tcp,
        ["Name", "Original Name", "Description", "Port"],
        [(s["name"], s["original_name"], s["description"], s.get("port", "")) for s in config["services"]["tcp"]]
    )

    # --- UDP Services ---
    ws_udp = wb.create_sheet("UDP Services")
    write_sheet(ws_udp,
        ["Name", "Original Name", "Description", "Port"],
        [(s["name"], s["original_name"], s["description"], s.get("port", "")) for s in config["services"]["udp"]]
    )

    # --- ICMP Services ---
    ws_icmp = wb.create_sheet("ICMP Services")
    write_sheet(ws_icmp,
        ["Name", "Original Name", "Description", "ICMP Version"],
        [(s["name"], s["original_name"], s["description"], s.get("icmp_version", "")) for s in config["services"]["icmp"]]
    )

    # --- Protocol Only ---
    ws_proto = wb.create_sheet("Protocol Only")
    write_sheet(ws_proto,
        ["Name", "Original Name", "Description", "Protocol", "Protocol Number"],
        [(s["name"], s["original_name"], s["description"], s.get("protocol", ""), s.get("protocol_number", "")) for s in config["services"]["protocol_only"]]
    )

    # --- Other Services ---
    ws_other = wb.create_sheet("Other Services")
    write_sheet(ws_other,
        ["Name", "Original Name", "Description", "Protocol"],
        [(s["name"], s["original_name"], s["description"], s.get("protocol", "")) for s in config["services"]["other"]]
    )

    # --- Service Groups ---
    ws_sg = wb.create_sheet("Service Groups")
    write_sheet(ws_sg,
        ["Name", "Original Name", "Description", "TCP/UDP Members", "ICMP Members", "Protocol Members", "Warnings"],
        [
            (
                g["name"], g["original_name"], g["description"],
                "; ".join(g.get("members", [])),
                "; ".join(g.get("icmp_members", [])),
                "; ".join(g.get("protocol_members", [])),
                "; ".join(g.get("warnings", []))
            )
            for g in config["service_groups"]
        ]
    )
    # Highlight warning cells
    for row_idx in range(2, len(config["service_groups"]) + 2):
        cell = ws_sg.cell(row=row_idx, column=7)
        if cell.value:
            cell.font = WARNING_FONT
            cell.fill = WARNING_FILL

    # --- Policies ---
    ws_pol = wb.create_sheet("Policies")
    write_sheet(ws_pol,
        ["Name", "Enabled", "Action", "Service", "Source Aliases", "Destination Aliases",
         "Source Members", "Destination Members", "NAT Policy", "Description",
         "Reject Action", "Tag", "Schedule", "Log Enabled", "Route Policy",
         "Proxy", "SDWAN Action", "App Action"],
        [
            (
                p["name"], p["enabled"], p["action"], p["service"],
                "; ".join(p["source_aliases"]),
                "; ".join(p["destination_aliases"]),
                "; ".join(p["source_members"]),
                "; ".join(p["destination_members"]),
                p["nat_policy"], p["description"], p["reject_action"],
                p["tag"], p["schedule"], p["log_enabled"],
                p["route_policy"], p["proxy"], p["sdwan_action"], p["app_action"]
            )
            for p in config["policies"]
        ]
    )

    # --- Policies (Resolved) ---
    # Same as Policies but source/destination resolved from alias names all the way
    # down to actual IP/network/FQDN values through the full alias -> group -> object chain
    resolver = build_address_resolver(config)
    ws_pol_res = wb.create_sheet("Policies (Resolved)")
    ws_pol_res.sheet_properties.tabColor = "7030A0"

    resolved_rows = []
    for p in config["policies"]:
        src_values = resolve_policy_values(p["source_aliases"], resolver)
        dst_values = resolve_policy_values(p["destination_aliases"], resolver)
        resolved_rows.append((
            p["name"], p["enabled"], p["action"], p["service"],
            "; ".join(p["source_aliases"]),
            "; ".join(p["destination_aliases"]),
            "; ".join(src_values),
            "; ".join(dst_values),
            p["nat_policy"], p["description"], p["reject_action"],
            p["tag"], p["schedule"], p["log_enabled"],
            p["route_policy"], p["proxy"], p["sdwan_action"], p["app_action"]
        ))

    write_sheet(ws_pol_res,
        ["Name", "Enabled", "Action", "Service", "Source Aliases", "Destination Aliases",
         "Source Values", "Destination Values", "NAT Policy", "Description",
         "Reject Action", "Tag", "Schedule", "Log Enabled", "Route Policy",
         "Proxy", "SDWAN Action", "App Action"],
        resolved_rows
    )

    # Highlight any unresolved references ([!] marker) in the resolved columns
    for row_idx in range(2, len(resolved_rows) + 2):
        for col in (7, 8):  # Source Values, Destination Values
            cell = ws_pol_res.cell(row=row_idx, column=col)
            if cell.value and "[!]" in str(cell.value):
                cell.font = WARNING_FONT
                cell.fill = WARNING_FILL

    # --- NAT Rules ---
    ws_nat = wb.create_sheet("NAT Rules")
    write_sheet(ws_nat,
        ["Name", "Type", "Algorithm", "Proxy ARP", "Address Type", "Port",
         "External Address", "Interface", "Internal Address"],
        [
            (
                n["name"], n["type"], n["algorithm"], n["proxy_arp"],
                n["address_type"], n["port"], n["external_address"],
                n["interface"], n["internal_address"]
            )
            for n in config["nat_rules"]
        ]
    )

    # --- Routes ---
    ws_routes = wb.create_sheet("Routes")
    write_sheet(ws_routes,
        ["Destination", "Mask", "Gateway"],
        [(r.get("destination", ""), r.get("mask", ""), r.get("gateway", "")) for r in config["routes"]]
    )

    # --- Interfaces ---
    ws_intf = wb.create_sheet("Interfaces")
    write_sheet(ws_intf,
        ["Name", "Description", "Device Name", "Enabled", "Node Type",
         "IP Address", "Mask", "Gateway", "Secondary IPs"],
        [
            (
                i["name"], i["description"], i["device_name"], i["enabled"],
                i["node_type"], i["ip"], i["mask"], i["gateway"],
                "; ".join(i["secondary_ips"])
            )
            for i in config["interfaces"]
        ]
    )

    # --- SDWAN Actions ---
    ws_sdwan = wb.create_sheet("SDWAN Actions")
    write_sheet(ws_sdwan,
        ["Name", "Description", "Algorithm", "Algorithm Description",
         "Primary Interface", "Secondary Interface", "All Interfaces",
         "Failback Grace Period"],
        [
            (
                a["name"], a["description"], a["algorithm"],
                a["algorithm_description"], a["primary_interface"],
                a["secondary_interface"], "; ".join(a["interfaces"]),
                a["failback_grace_period"]
            )
            for a in config["sdwan_actions"]
        ]
    )

    # --- App Actions ---
    ws_app = wb.create_sheet("App Actions")
    write_sheet(ws_app,
        ["Name", "Description", "Fallthrough Action", "Allowed Apps", "Blocked Apps"],
        [
            (
                a["name"], a["description"], a["fallthrough_action"],
                "; ".join(a["allowed_apps"]),
                "; ".join(a["blocked_apps"])
            )
            for a in config["app_actions"]
        ]
    )

    # --- Geo Actions ---
    ws_geo = wb.create_sheet("Geo Actions")
    write_sheet(ws_geo,
        ["Name", "Description", "Blocked Countries"],
        [(g["name"], g["description"], "; ".join(g["blocked_countries"])) for g in config["geo_actions"]]
    )

    # --- Validation ---
    ws_val = wb.create_sheet("Validation")
    ws_val.sheet_properties.tabColor = "CC0000" if (issues["broken_address_references"] or issues["broken_alias_references"]) else "00B050"
    val_rows = []
    for issue in issues["broken_address_references"]:
        val_rows.append(("Address Reference", issue["group"], issue["missing_member"]))
    for issue in issues["broken_alias_references"]:
        val_rows.append(("Alias Reference", issue["group"], issue["missing_alias"]))
    if not val_rows:
        val_rows.append(("No issues found", "", ""))

    write_sheet(ws_val,
        ["Issue Type", "Group", "Missing Reference"],
        val_rows
    )
    # Highlight validation issues
    for row_idx in range(2, len(val_rows) + 2):
        for col in range(1, 4):
            cell = ws_val.cell(row=row_idx, column=col)
            if cell.value and cell.value != "No issues found":
                cell.font = WARNING_FONT

    wb.save(output_file)
    return wb


# ============================================================================
# Main
# ============================================================================

if __name__ == "__main__":
    if len(sys.argv) < 2:
        from pathlib import Path
        print(f"Usage: ./{Path(sys.argv[0]).name} <config.xml>")
        sys.exit(1)

    xml_file = sys.argv[1]

    config = parse_watchguard_config(xml_file)
    issues = validate_references(config)

    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M')
    xlsx_file = f"watchguard_config_v5_{timestamp}.xlsx"

    write_workbook(config, issues, xlsx_file)

    # Print summary to console
    print(f"[OK] Parsed configuration from {xml_file}")
    print(f"[OK] Output written to {xlsx_file}")
    print(f"\n{'='*60}")
    print("ADDRESS OBJECTS")
    print(f"{'='*60}")
    print(f"Hosts:             {len(config['addresses']['hosts'])}")
    print(f"Networks:          {len(config['addresses']['networks'])}")
    print(f"Ranges:            {len(config['addresses']['ranges'])}")
    print(f"FQDNs:             {len(config['addresses']['fqdns'])}")
    print(f"Address Groups:    {len(config['address_groups'])}")
    print(f"Interface Aliases: {len(config['interface_aliases'])}")
    print(f"\n{'='*60}")
    print("SERVICE OBJECTS")
    print(f"{'='*60}")
    print(f"TCP Services:      {len(config['services']['tcp'])}")
    print(f"UDP Services:      {len(config['services']['udp'])}")
    print(f"ICMP Services:     {len(config['services']['icmp'])}")
    print(f"Protocol-Only:     {len(config['services']['protocol_only'])}")
    print(f"Other/Unsupported: {len(config['services']['other'])}")
    print(f"Service Groups:    {len(config['service_groups'])}")
    print(f"\n{'='*60}")
    print("POLICIES & OTHER")
    print(f"{'='*60}")
    print(f"Policies:          {len(config['policies'])}")
    print(f"NAT Rules:         {len(config['nat_rules'])}")
    print(f"Routes:            {len(config['routes'])}")
    print(f"Interfaces:        {len(config['interfaces'])}")
    print(f"SDWAN Actions:     {len(config['sdwan_actions'])}")
    print(f"App Actions:       {len(config['app_actions'])}")
    print(f"Geo Actions:       {len(config['geo_actions'])}")

    if issues["broken_address_references"]:
        print(f"\n[!] {len(issues['broken_address_references'])} broken address references (see Validation sheet)")
    if issues["broken_alias_references"]:
        print(f"\n[!] {len(issues['broken_alias_references'])} broken alias references (see Validation sheet)")
